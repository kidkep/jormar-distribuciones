import io
import zipfile
from datetime import datetime

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncConnection
from openpyxl import Workbook

import app.models  # noqa: F401  (registra todas las tablas en Base.metadata)


def _serialize(value):
    if value is None:
        return None
    if isinstance(value, (datetime,)):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)


def _escape_sql(value):
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


async def extract_tables(conn: AsyncConnection) -> dict[str, list[list]]:
    """Devuelve {nombre_tabla: [filas como listas]} para todas las tablas de la BD."""
    from app.database import Base

    tables: dict[str, list[list]] = {}
    for table in Base.metadata.sorted_tables:
        result = await conn.execute(text(f'SELECT * FROM "{table.name}"'))
        rows = result.fetchall()
        columns = list(result.keys())
        data = [list(columns)]
        for row in rows:
            data.append([_serialize(v) for v in row])
        tables[table.name] = data
    return tables


def build_xlsx_bytes(tables: dict[str, list[list]]) -> bytes:
    """Genera un archivo Excel en memoria, una hoja por tabla."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, data in tables.items():
        ws = wb.create_sheet(title=name[:31])
        if data:
            ws.append(data[0])
            for row in data[1:]:
                ws.append(row)
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def build_sql_dump(conn: AsyncConnection, tables: dict[str, list[list]]) -> str:
    """Genera un dump SQL restaurable: CREATE TABLE + INSERT INTO por cada tabla."""
    from app.database import Base

    lines: list[str] = []
    lines.append("-- JORMAR DISTRIBUCIONES - BACKUP DE BASE DE DATOS")
    lines.append(f"-- Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")

    ddl = {}
    for table in Base.metadata.sorted_tables:
        ddl[table.name] = str(str(table.compile(dialect=conn.dialect)).strip())

    for table in Base.metadata.sorted_tables:
        name = table.name
        if name in ddl:
            lines.append(f'DROP TABLE IF EXISTS "{name}" CASCADE;')
            lines.append(ddl[name].rstrip().rstrip(";") + ";")
            lines.append("")

        data = tables.get(name, [])
        if len(data) > 1:
            columns = data[0]
            for row in data[1:]:
                cols = ", ".join(f'"{c}"' for c in columns)
                vals = ", ".join(_escape_sql(v) for v in row)
                lines.append(f'INSERT INTO "{name}" ({cols}) VALUES ({vals});')
            lines.append("")

    return "\n".join(lines)


def build_backup_zip_bytes(xlsx_bytes: bytes, sql_text: str, fecha: str) -> bytes:
    """Empaqueta el Excel y el SQL en un ZIP en memoria."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"base_de_datos_{fecha}.xlsx", xlsx_bytes)
        zf.writestr(f"backup_{fecha}.sql", sql_text)
    buf.seek(0)
    return buf.getvalue()
