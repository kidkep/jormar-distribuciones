import requests
import openpyxl
import re

EXCEL = r"C:\Users\JUANCA\Desktop\Jormar\P2 GESTOR DE NEGOCIO ORIGINAL No 1   2026 (1) (1).xlsm"
BASE = "https://jormar-api.onrender.com/api/v1"

EXCLUIR_PROVEEDORES = {"gasto total camioneta", "variados", "varios"}


def login():
    r = requests.post(f"{BASE}/auth/login", json={"username": "CEO", "password": "2908"})
    r.raise_for_status()
    return r.json()["access_token"]


def clean_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def to_digits(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(int(value)) if int(value) != 0 else None
    nums = re.findall(r"\d+", str(value))
    return nums[0] if nums else None


def normalize_doc(value):
    """Devuelve solo digitos del documento, o None si no tiene digitos validos."""
    if value is None:
        return None
    nums = re.findall(r"\d+", str(value))
    if not nums:
        return None
    joined = "".join(nums)
    if joined in ("0",):
        return None
    return joined


def to_number(value, default=0):
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    nums = re.findall(r"[\d.,]+", str(value))
    if not nums:
        return default
    try:
        return float(nums[0].replace(",", ""))
    except ValueError:
        return default


token = login()
headers = {"Authorization": f"Bearer {token}"}
print("LOGIN OK")

wb = openpyxl.load_workbook(EXCEL, read_only=True, keep_vba=True)

# 1. IMPORT SUPPLIERS
print("\n--- IMPORTING SUPPLIERS ---")
ws = wb["BD_PROVEEDORES"]
suppliers_ok = 0
suppliers_fail = 0
skipped = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    name = clean_text(row[0])
    if not name or name.strip().lower() in EXCLUIR_PROVEEDORES:
        skipped += 1
        continue
    nit = normalize_doc(row[2])
    phone = to_digits(row[3])
    address = clean_text(row[4])
    email = clean_text(row[5])
    city = clean_text(row[6])

    if not nit or nit.lower() in ("impuestos", "xxxx", "xxxxxxxxxx"):
        nit = f"PROV-{row_idx}"

    data = {
        "name": name,
        "document_type": "NIT",
        "document_number": nit,
        "phone": phone,
        "address": address,
        "email": email,
        "city": city,
    }

    r = requests.post(f"{BASE}/suppliers", headers=headers, json=data)
    if r.status_code == 201:
        suppliers_ok += 1
    elif r.status_code == 409:
        # ya existe (documento ocupado por este u otro proveedor) -> omitir
        pass
    else:
        suppliers_fail += 1
        if suppliers_fail <= 5:
            print(f"  FAIL: {name} -> {r.status_code}: {r.text[:200]}")

print(f"  Suppliers: {suppliers_ok} created, {suppliers_fail} errors, {skipped} skipped")

# 2. IMPORT CLIENTS
print("\n--- IMPORTING CLIENTS ---")
ws = wb["BD_CLIENTES"]
clients_ok = 0
clients_fail = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    name = clean_text(row[0])
    if not name:
        continue
    doc = normalize_doc(row[1])
    phone = to_digits(row[2])
    address = clean_text(row[3])
    email = clean_text(row[4])
    city = clean_text(row[5])

    if not doc or doc.lower() in ("xxxxxx", "impuestos", "xxxx"):
        doc = f"CLI-{row_idx}"

    data = {
        "name": name,
        "document_type": "CC",
        "document_number": doc,
        "phone": phone,
        "address": address,
        "email": email,
        "city": city,
    }

    r = requests.post(f"{BASE}/clients", headers=headers, json=data)
    if r.status_code == 201:
        clients_ok += 1
    elif r.status_code == 409:
        # ya existe (documento ocupado) -> omitir
        pass
    else:
        clients_fail += 1
        if clients_fail <= 5:
            print(f"  FAIL: {name} -> {r.status_code}: {r.text[:200]}")

print(f"  Clients: {clients_ok} created, {clients_fail} errors")

# 3. IMPORT PRODUCTS
print("\n--- IMPORTING PRODUCTS ---")
ws = wb["BD_PRODUCTOS"]
products_ok = 0
products_fail = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    code = row[0]
    name = clean_text(row[1])
    purchase = to_number(row[2])
    sale = to_number(row[3])
    stock = row[4]

    if not name:
        continue

    sku = clean_text(str(code)) if code is not None else None
    if not sku:
        sku = f"PROD-{row_idx}"

    data = {
        "sku": sku,
        "name": name,
        "purchase_price": purchase,
        "sale_price": sale,
        "current_stock": int(stock) if stock and isinstance(stock, (int, float)) else 0,
        "min_stock": 5,
        "tax_rate": 19,
        "is_active": True,
    }

    r = requests.post(f"{BASE}/products", headers=headers, json=data)
    if r.status_code == 201:
        products_ok += 1
    elif r.status_code == 409:
        # ya existe -> omitir
        pass
    else:
        products_fail += 1
        if products_fail <= 5:
            print(f"  FAIL: {name} -> {r.status_code}: {r.text[:200]}")

print(f"  Products: {products_ok} created, {products_fail} errors")

# SUMMARY
print("\n=== IMPORT SUMMARY ===")
print(f"  Suppliers: {suppliers_ok}")
print(f"  Clients:   {clients_ok}")
print(f"  Products:  {products_ok}")
print("\nDone!")
