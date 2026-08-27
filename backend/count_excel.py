import openpyxl

EXCEL = r"C:\Users\JUANCA\Desktop\JORMAR\P2 GESTOR DE NEGOCIO ORIGINAL No 1   2026 (1) (1).xlsm"
wb = openpyxl.load_workbook(EXCEL, read_only=True, keep_vba=True)

for name in ["BD_CLIENTES", "BD_PRODUCTOS", "BD_PROVEEDORES", "BD_GASTOSYCOSTOS", "KARDEX", "DEUDORES", "DINEROS"]:
    ws = wb[name]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row[:7]):
            count += 1
    print(f"{name}: {count} data rows")

# Show full BD_PRODUCTOS first few + last
ws = wb["BD_PRODUCTOS"]
rows = list(ws.iter_rows(min_row=2, values_only=True))
valid = [r for r in rows if r[0] is not None and r[1] is not None]
print(f"\nBD_PRODUCTS unique: {len(valid)}")
print(f"  First: {valid[0][:4]}")
print(f"  Last: {valid[-1][:4]}")

# Show full BD_CLIENTES count
ws = wb["BD_CLIENTES"]
rows = list(ws.iter_rows(min_row=2, values_only=True))
valid = [r for r in rows if r[0] is not None]
print(f"\nBD_CLIENTES unique: {len(valid)}")

ws = wb["BD_PROVEEDORES"]
rows = list(ws.iter_rows(min_row=2, values_only=True))
valid = [r for r in rows if r[0] is not None]
print(f"BD_PROVEEDORES unique: {len(valid)}")

ws = wb["BD_GASTOSYCOSTOS"]
rows = list(ws.iter_rows(min_row=2, values_only=True))
valid = [r for r in rows if r[0] is not None]
print(f"BD_GASTOSYCOSTOS unique: {len(valid)}")
