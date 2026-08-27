import openpyxl

EXCEL = r"C:\Users\JUANCA\Desktop\JORMAR\P2 GESTOR DE NEGOCIO ORIGINAL No 1   2026 (1) (1).xlsm"
wb = openpyxl.load_workbook(EXCEL, read_only=True, keep_vba=True)
print("Sheets:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(max_row=5, values_only=True))
    print(f"\n=== {name} ===")
    for i, row in enumerate(rows):
        print(f"  Row {i}: {row}")
