import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_PATH = os.path.join(os.path.dirname(__file__), "backend", "jormar.db")
EXPORTS_DIR = os.path.join(os.path.dirname(__file__), "exportaciones")

TABLES = [
    ("users", "Usuarios"),
    ("roles", "Roles"),
    ("clients", "Clientes"),
    ("products", "Productos"),
    ("suppliers", "Proveedores"),
    ("categories", "Categorias"),
    ("units", "Unidades"),
    ("sales", "Ventas"),
    ("sale_items", "Detalle Ventas"),
    ("quotes", "Cotizaciones"),
    ("quote_items", "Detalle Cotizaciones"),
    ("expenses", "Gastos"),
    ("payments", "Pagos"),
]

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def exportar():
    os.makedirs(EXPORTS_DIR, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)

    for table_name, sheet_name in TABLES:
        try:
            cursor.execute(f"SELECT * FROM {table_name}")
            rows = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
        except Exception:
            continue

        ws = wb.create_sheet(title=sheet_name)

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER

        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(str(col_name))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    conn.close()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"jormar_exportacion_{timestamp}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)
    wb.save(filepath)
    print(f"Exportacion creada: {filepath}")


if __name__ == "__main__":
    exportar()
